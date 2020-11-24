
import yaml
import openpyxl
from itertools import islice

row_elements = 'username firstname lastname group examid'.split(' ')

def noop(*args, **kwargs): pass
def lmap(*args, **kwargs): return list(map(*args, **kwargs))
def at(k):                 return lambda o: o[k]
def attr_value():          return lambda o: None if o is None else o.value
def ofdict(o):             return lambda k: None if k not in o else o[k]
def oflist(o):             return lambda k: None if k is None else o[k]

def read_config(pro, filename='parasite.yaml', print=noop):
    with open(pro + '/' + filename) as f:
        cfg = yaml.full_load(f)
    def withdef(path, v, o=cfg, prev=''):
        if '/' in path:
            splt = path.split('/', 1)
            if splt[0] not in o:
                o[splt[0]] = {}
            withdef(splt[1], v, o[splt[0]], prev+'→'+splt[0])
        else:
            if path not in o:
                print("At", prev+'→'+path, "--- Setting default", v)
                o[path] = v
            else:
                print("At", prev+'→'+path, "--- Keeping", o[path], "vs default", v)
    for h in row_elements:
        withdef('headers/'+h, h)
    return cfg

def parse_xlsx_generic(xlfile, sheet_index=0, sheet_name=None):
    wb = openpyxl.load_workbook(filename = xlfile)
    if sheet_name is None:
        sheet_name = wb.sheetnames[sheet_index] 
    ws = wb[sheet_name]
    print(wb.sheetnames, sheet_index, sheet_name, ws)
    res = []
    for i, row in enumerate(ws.rows):
        res.append(lmap(attr_value(), row))
    return res

def add_xlsx_sheet_keyed_by_examid(xlfile, data):
    header = data['header']   # header[f]
    content = data['content'] # content[f][u]

    wb = openpyxl.load_workbook(filename = xlfile)

    try:
        stitle = 1
        title = data['title'] % (stitle,)
        format_title = True
    except:
        title = data['title']
        format_title = False

    if format_title:
        for ititle in range(stitle, 1000): # force stop
            title = data['title'] % (ititle,)
            if title not in wb.sheetnames:
                break

    ws = wb.create_sheet(title=title)
    ws.cell(row=1, column=1, value='examid')
    for f, h in enumerate(header):
        ws.cell(row=1, column=2+f, value=h)

    users = sorted(list(set().union(*[[int(k) for k in mapp.keys()] for mapp in content])))
    print(users)

    print(content)

    for i, u_int in enumerate(users):
        u = str(u_int)
        ws.cell(row=2+i, column=1, value=u)
        for f, h in enumerate(header):
            if f < len(content) and u in content[f]:
                print(i,u,f,h,content[f][u])
                ws.cell(row=2+i, column=2+f, value=content[f][u])

    wb.save(xlfile)


def parse_xlsx_first_page(xlfile, data, header_names, log):
    write = None
    should_save = False
    if 'write' in data: # supposing first and last names match (so already there)
        write = [[None, None, None, None, None]] + data['write'] # add one pseudo line to keep the row id of xls
        should_save = True

    wb = openpyxl.load_workbook(filename = xlfile)
    ws = wb[wb.sheetnames[0]]
    def digest_header_into_index_access():
        headers = lmap(attr_value(), ws['1'])
        headers_lookup = { v: i for i,v in enumerate(headers) }
        cfg_mapped = lmap(lambda h: header_names[h], row_elements)
        return lmap(ofdict(headers_lookup), cfg_mapped)

    row_indices = digest_header_into_index_access()
    res = []
    for i, row in islice(enumerate(ws.rows), 1, None):
        cells = lmap(oflist(row), row_indices)
        if write is not None:
            if write[i][1:3] == lmap(attr_value(), cells[1:3]):
                for j in range(3, min(len(cells), len(write))):
                    if cells[j] is not None:
                        cells[j].value = write[i][j]
            else:
                log('alert', 'Not matching ['+str(i)+']'+str(lmap(attr_value(), cells))+' VS '+str(write[i]))
        res.append(lmap(attr_value(), cells))

    if should_save:
        wb.save(xlfile)

    return res
