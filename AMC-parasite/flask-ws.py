
from flask import Flask, render_template, send_from_directory
from flask_socketio import SocketIO, join_room, emit

import time
import os
import shutil
import codecs
import openpyxl
from itertools import islice

from tools_querysqlite import *
from tools_usetorch import *
from tools_generic import *

# pip3 install flask flask-socketio eventlet
# https://secdevops.ai/weekend-project-part-2-turning-flask-into-a-real-time-websocket-server-using-flask-socketio-ab6b45f1d896
# pip3 install openpyxl yaml

# initialize Flask
#app = Flask(__name__, template_folder='flask-ws')
app = Flask(__name__)
socketio = SocketIO(app)

local_MC = './,,test/'

# Host the mc projects
@app.route('/MC/<path:path>')
def send_MC(path):
    return send_from_directory(local_MC, path)
# Host the generate files
@app.route('/gen/<path:path>')
def send_publiccustom(path):
    return send_from_directory('./generated/', path)

@app.route('/')
def index():
    """Serve the index HTML"""
    return render_template('plain-index.html')





@socketio.on('manual-load-images')
def on_manual_load_images(data):
    predict = 'predict' in data
    print('SQL QUERY')
    if 'only' in data:
        info = preload_all_queries(make_connection(local_MC + data['pro'] + '/data/capture.sqlite'), more=' AND student='+str(data['only']), improcess=assuch)
    else:
        info = preload_all_queries(make_connection(local_MC + data['pro'] + '/data/capture.sqlite'), improcess=assuch)
	# 0id_answer 1student
    # 2`zoneid`	INTEGER, 3`student`	INTEGER, 4`page`	INTEGER, 5`copy`	INTEGER, 6`type`	INTEGER, 7`id_a`	INTEGER, 8`id_b`	INTEGER,
    # 9`total`	INTEGER DEFAULT -1, 10`black`	INTEGER DEFAULT -1, 11`manual`	REAL DEFAULT -1, 12`image`	TEXT, 13`imagedata`	BLOB,
    if predict:
        print("Loading pytorch model")
        net = torch.load('resources/model-emnist4big.torch').to(torch.device('cpu'))
    print('...DONE')
    sub = 'pyannotate'
    directory = './generated/'+sub
    if os.path.exists(directory):
        shutil.rmtree(directory)
    if not os.path.exists(directory):
        os.makedirs(directory)
    for k,v in info.items():
        pairs = list(enumerate(v))
        ims = []
        for i,r in pairs:
            n = "/im-" + str(i) + ".png"
            #print('IMAGE', n)
            #print(info[k][:-2])
            #info[k].append(sub+n)
            #print(i, n)
            with open(directory+n, "wb") as out_file:
                out_file.write(r[-1])
                v[i] = r[:-1] + ('gen/'+sub+n,)
            if predict:
                ims.append(bytes2im(r[-1]))
        if predict:
            with torch.no_grad():
                pred = net(torch.from_numpy(np.array(ims)))
            amax = np.argmax(pred.numpy(), axis=1)
            our_classes = "=:;.,-_()[]!?*/'"
            classes = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabdefghnqrt"+our_classes
            for i,r in pairs:
                ch = classes[int(amax[i])]
                if r[10] == 0: ch = '_'
                v[i] = v[i][:-2] + (ch,) + v[i][-1:]
                #print(v[i])

    info['_id'] = data['_id']
    emit('manual-loaded-images', info)

@socketio.on('manual-log')
def on_manual_log(data):
    logfile = local_MC + data['pro'] + '/parasite-logs-manual.jstream'
    with codecs.open(logfile, "a", "utf-8") as f:
        f.write(data['data'])
    print("saved")

@socketio.on('miniset-get-logs')
def on_miniset_getlogs(data):
    logfile = local_MC + data['pro'] + '/parasite-logs-manual.jstream'
    with codecs.open(logfile, "r", "utf-8") as f:
        emit('miniset-got-logs', list(map(lambda l: l[:-1], f.readlines())))

@socketio.on('miniset-export')
def on_miniset_export(data):
    name = data['name']
    print("Exporting miniset", name)
    sub = 'miniset/'+name
    directory = 'generated/'+sub
    if not os.path.exists(directory):
        os.makedirs(directory)
    else:
        print("WILL NOT OVERWRITE", directory)
        return
    for d in data['annotations']:
        student = d[0]
        ann = d[1]
        info = preload_all_queries(make_connection(local_MC + data['pro'] + '/data/capture.sqlite'), more=' AND student='+str(student), improcess=assuch)
        info = info[int(student)]
        for i, r in list(enumerate(info)):
            if not str(i) in ann:
                print("SKIP", i)
                continue
            if ann[str(i)] == ' ': continue  # skip space annotated images
            di = directory + '/class-' + ann[str(i)]
            if not os.path.exists(di):
                os.makedirs(di)
            n = '/im-%05d.png' % (i,)
            with open(di+n, "wb") as out_file:
                out_file.write(r[-1])
            #n = '/im-%05d.txt' % (i,)
            #with open(directory+n, "w") as out_file:
            #    out_file.write(ann[str(i)])
    print("WROTE", directory)


@socketio.on('xlsx-structured-rows')
def on_xlsx_read_rows(data):
    cfg = read_config(local_MC + data['pro'])
    xlfile = local_MC + data['pro'] + '/parasite.xlsx'
    wb = openpyxl.load_workbook(filename = xlfile)
    ws = wb[wb.sheetnames[0]]
    def digest_header_into_index_access():
        headers = lmap(attr_value(), ws['1'])
        headers_lookup = { v: i for i,v in enumerate(headers) }
        cfg_mapped = lmap(lambda h: cfg['headers'][h], row_elements)
        return lmap(of(headers_lookup), cfg_mapped)

    row_indices = digest_header_into_index_access()
    res = []
    for i, row in islice(enumerate(ws.rows), 1, None):
        cells = lmap(of(row), row_indices)
        res.append(lmap(attr_value(), cells))
    emit('got-xlsx-structured-rows', res)


if __name__ == '__main__':
    socketio.run(app, debug=True)
