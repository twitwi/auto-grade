
from flask import Flask, Blueprint, render_template, send_from_directory
from flask_socketio import SocketIO, join_room, emit

import time
import os
import shutil
import codecs
import openpyxl
from itertools import islice

from tools_querysqlite import *
from tools_usetorch import *
from tools_generic import *

from flask_app import socketio

@socketio.on('xlsx-structured-rows')
def on_xlsx_read_rows(data):
    cfg = read_config(local_MC + data['pro'])
    xlfile = local_MC + data['pro'] + '/parasite.xlsx'

    write = None
    should_save = False
    if 'write' in data: # supposing first and last names match (so already there)
        write = [[None, None, None, None, None]] + data['write'] # add one pseudo line to keep the row id of xls
        should_save = True

    wb = openpyxl.load_workbook(filename = xlfile)
    ws = wb[wb.sheetnames[0]]
    def digest_header_into_index_access():
        headers = lmap(attr_value(), ws['1'])
        headers_lookup = { v: i for i,v in enumerate(headers) }
        cfg_mapped = lmap(lambda h: cfg['headers'][h], row_elements)
        return lmap(ofdict(headers_lookup), cfg_mapped)

    row_indices = digest_header_into_index_access()
    res = []
    for i, row in islice(enumerate(ws.rows), 1, None):
        cells = lmap(oflist(row), row_indices)
        if write is not None:
            if write[i][1:3] == lmap(attr_value(), cells[1:3]):
                for j in range(3, min(len(cells), len(write))):
                    if cells[j] is not None:
                        cells[j].value = write[i][j]
            else:
                emit('alert', 'Not matching ['+str(i)+']'+str(lmap(attr_value(), cells))+' VS '+str(write[i]))
        res.append(lmap(attr_value(), cells))

    if should_save:
        wb.save(xlfile)
    emit('got-xlsx-structured-rows' if 'callback' not in data else data['callback'], res)

@socketio.on('xlsx-add-sheet')
def on_xlsx_read_rows(data):
    from collections import Set
    cfg = read_config(local_MC + data['pro'])
    xlfile = local_MC + data['pro'] + '/parasite.xlsx'

    head = data['head'] # head[f]
    content = data['content'] # content[f][u]

    wb = openpyxl.load_workbook(filename = xlfile)

    try:
        stitle = 0
        title = data['title'] % (stitle,)
        format_title = True
    except:
        title = data['title']
        format_title = False

    if format_title:
        for ititle in range(stitle, 1000): # force stop
            title = data['title'] % (ititle,)
            if title not in wb.sheetnames:
                break

    ws = wb.create_sheet(title=title)
    ws.cell(row=1, column=1, value='examid')
    for f, h in enumerate(head):
        ws.cell(row=1, column=2+f, value=h)

    users = sorted(list(set().union(*[[int(k) for k in map.keys()] for map in content])))
    print(users)

    print(content)

    for i, u_int in enumerate(users):
        u = str(u_int)
        ws.cell(row=2+i, column=1, value=u)
        for f, h in enumerate(head):
            if f < len(content) and u in content[f]:
                print(i,u,f,h,content[f][u])
                ws.cell(row=2+i, column=2+f, value=content[f][u])

    wb.save(xlfile)
    emit('xlsx-added-sheet' if 'callback' not in data else data['callback'])

@socketio.on('yaml-config')
def on_get_yaml_config(data):
    cfg = read_config(local_MC + data['pro'])
    emit('got-yaml-config', cfg)
